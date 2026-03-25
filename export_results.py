import pandas as pd
import numpy as np
import semopy
from scipy import stats
from sklearn.decomposition import PCA
from linearmodels.iv import IV2SLS
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import zipfile, os, warnings
warnings.filterwarnings('ignore')

plt.rcParams['font.family'] = 'DejaVu Sans'
OUT = 'SEM_PathAnalysis_Results#6'
os.makedirs(OUT, exist_ok=True)

# ── 데이터 ────────────────────────────────────────────────────
df_raw = pd.read_excel('E:/data/aesop2025/aesop2025.xlsx', sheet_name='시군구(행정구제외)')
df_raw = df_raw[df_raw['basic_gov'].notna()].copy()

vars_disparity = ['bnbl_rate', 'oldb']                   # Urban Disparity (빈곤 + 노후화)
vars_dev       = ['lit_pc', 'grdp_pc', 'hale']           # Urban Development Level (통제변수)
vars_uci       = ['DD', 'LUM', 'AC', 'SA', 'subs']      # Urban Spatial Structure (+ 도시철도 유무)
vars_carbon    = ['build_elec_e_pc', 'transport_e_pc']
all_vars       = vars_disparity + vars_dev + vars_uci + vars_carbon

df_model = df_raw[all_vars].dropna().copy()
df_z = df_model.copy()
for col in all_vars:
    df_z[col] = stats.zscore(df_model[col], nan_policy='omit')

N = len(df_model)
print(f"분석 관측치: {N}개")


# ════════════════════════════════════════════════════════════════
# 1. Non-recursive SEM
# ════════════════════════════════════════════════════════════════
model_nr = semopy.Model("""
    urban_disparity =~ bnbl_rate + oldb
    urban_dev       =~ lit_pc + grdp_pc + hale
    UCI             =~ DD + LUM + AC + SA + subs
    carbon          =~ build_elec_e_pc + transport_e_pc

    UCI             ~ urban_disparity + urban_dev
    urban_disparity ~ UCI + urban_dev
    carbon          ~ UCI + urban_disparity + urban_dev

    UCI             ~~ urban_disparity
""")
model_nr.fit(df_z)
params_nr = model_nr.inspect()
fit_nr    = semopy.calc_stats(model_nr)

latent_vars = ['urban_disparity', 'urban_dev', 'UCI', 'carbon']
structural_mask  = (params_nr['op'] == '~') & \
                   params_nr['lval'].isin(latent_vars) & \
                   params_nr['rval'].isin(latent_vars)
measurement_mask = (params_nr['op'] == '~') & \
                   params_nr['rval'].isin(latent_vars) & \
                   ~params_nr['lval'].isin(latent_vars)

est = params_nr.set_index(['lval', 'op', 'rval'])['Estimate']
se  = params_nr.set_index(['lval', 'op', 'rval'])['Std. Err']
pv  = params_nr.set_index(['lval', 'op', 'rval'])['p-value']

a            = est.get(('UCI',             '~', 'urban_disparity'), np.nan)
b            = est.get(('carbon',          '~', 'UCI'),             np.nan)
c            = est.get(('carbon',          '~', 'urban_disparity'), np.nan)
d            = est.get(('urban_disparity', '~', 'UCI'),             np.nan)
e_dev_uci    = est.get(('UCI',             '~', 'urban_dev'),       np.nan)
e_dev_disp   = est.get(('urban_disparity', '~', 'urban_dev'),       np.nan)
e_dev_carbon = est.get(('carbon',          '~', 'urban_dev'),       np.nan)

a_p          = pv.get(('UCI',             '~', 'urban_disparity'), np.nan)
b_p          = pv.get(('carbon',          '~', 'UCI'),             np.nan)
c_p          = pv.get(('carbon',          '~', 'urban_disparity'), np.nan)
d_p          = pv.get(('urban_disparity', '~', 'UCI'),             np.nan)
e_dev_uci_p  = pv.get(('UCI',             '~', 'urban_dev'),       np.nan)
e_dev_disp_p = pv.get(('urban_disparity', '~', 'urban_dev'),       np.nan)
e_dev_carb_p = pv.get(('carbon',          '~', 'urban_dev'),       np.nan)

indirect = a * b
total    = c + indirect
dev_indirect = e_dev_uci * b
dev_total    = e_dev_carbon + dev_indirect


# ════════════════════════════════════════════════════════════════
# 2. 2SLS
# ════════════════════════════════════════════════════════════════
df_z2 = df_z.copy().reset_index(drop=True)

def get_pca_score(data, cols):
    pca = PCA(n_components=1)
    return pca.fit_transform(stats.zscore(data[cols].values, axis=0)).flatten()

df_z2['disparity_score'] = get_pca_score(df_z2, vars_disparity)
df_z2['uci_score']       = get_pca_score(df_z2, vars_uci)
df_z2['dev_score']       = get_pca_score(df_z2, vars_dev)

iv_results = {}
for cv in vars_carbon:
    res = IV2SLS(
        dependent   = df_z2[cv],
        exog        = df_z2[['dev_score']],
        endog       = df_z2[['uci_score', 'disparity_score']],
        instruments = df_z2[vars_uci + vars_disparity]
    ).fit(cov_type='robust')
    iv_results[cv] = res


# ════════════════════════════════════════════════════════════════
# 공통 헬퍼
# ════════════════════════════════════════════════════════════════
def sig_star(p):
    if pd.isna(p): return ''
    try: p = float(p)
    except (ValueError, TypeError): return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return ''

def safe_round(val, n=4):
    try: return round(float(val), n)
    except (ValueError, TypeError): return ''

def style_header(ws, row, cols, color='1F4E79'):
    fill = PatternFill('solid', fgColor=color)
    font = Font(color='FFFFFF', bold=True, size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')

def border_range(ws, r1, c1, r2, c2):
    thin = Side(style='thin')
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ════════════════════════════════════════════════════════════════
# 엑셀 저장
# ════════════════════════════════════════════════════════════════
wb = Workbook()
headers = ['LHS', 'Op', 'RHS', 'Estimate', 'Std. Err', 'z-value', 'p-value', 'Sig.']

# ── Sheet 1: SEM 구조 경로 ───────────────────────────────────
ws1 = wb.active
ws1.title = 'SEM_Structural'
for col, w in zip(['A','B','C','D','E','F','G','H'], [22,6,22,13,13,13,13,8]):
    ws1.column_dimensions[col].width = w

ws1['A1'] = 'Non-recursive SEM — Structural Paths'
ws1['A1'].font = Font(bold=True, size=12)
ws1.merge_cells('A1:H1')
ws1['A2'] = (
    'urban_disparity (bnbl_rate, no_subs, oldb) <-> UCI (DD,LUM,AC,SA)  |  '
    'urban_dev (lit_pc, grdp_pc, hale) = exogenous control'
)
ws1['A2'].font = Font(italic=True, size=9, color='555555')
ws1.merge_cells('A2:H2')

for i, h in enumerate(headers, 1):
    ws1.cell(3, i, h)
style_header(ws1, 3, len(headers))

structural_rows = params_nr[structural_mask].reset_index(drop=True)
for r_i, row in structural_rows.iterrows():
    p = row['p-value']
    data = [row['lval'], row['op'], row['rval'],
            safe_round(row['Estimate']), safe_round(row['Std. Err']),
            safe_round(row['z-value'], 3), safe_round(p), sig_star(p)]
    for c_i, val in enumerate(data, 1):
        ws1.cell(r_i + 4, c_i, val).alignment = Alignment(horizontal='center')
border_range(ws1, 3, 1, len(structural_rows) + 4, len(headers))

# ── Sheet 2: SEM 측정 모델 ───────────────────────────────────
ws2 = wb.create_sheet('SEM_Measurement')
for col, w in zip(['A','B','C','D','E','F','G','H'], [22,6,22,13,13,13,13,8]):
    ws2.column_dimensions[col].width = w

ws2['A1'] = 'Non-recursive SEM — Measurement Model (4 Constructs)'
ws2['A1'].font = Font(bold=True, size=12)
ws2.merge_cells('A1:H1')
for i, h in enumerate(headers, 1):
    ws2.cell(3, i, h)
style_header(ws2, 3, len(headers), color='2E7539')

measurement_rows = params_nr[measurement_mask].reset_index(drop=True)
for r_i, row in measurement_rows.iterrows():
    p = row['p-value']
    data = [row['lval'], row['op'], row['rval'],
            safe_round(row['Estimate']), safe_round(row['Std. Err']),
            safe_round(row['z-value'], 3), safe_round(p), sig_star(p)]
    for c_i, val in enumerate(data, 1):
        ws2.cell(r_i + 4, c_i, val).alignment = Alignment(horizontal='center')
border_range(ws2, 3, 1, len(measurement_rows) + 4, len(headers))

# ── Sheet 3: 모델 적합도 ─────────────────────────────────────
ws3 = wb.create_sheet('Model_Fit')
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 25

ws3['A1'] = 'Model Fit Indices'
ws3['A1'].font = Font(bold=True, size=12)
ws3.merge_cells('A1:C1')
for i, h in enumerate(['Index', 'Value', 'Threshold'], 1):
    ws3.cell(3, i, h)
style_header(ws3, 3, 3, color='7B3F00')

thresholds = {
    'DoF': '', 'chi2': '', 'chi2 p-value': '> .05 (good fit)',
    'CFI': '> .95 (good)', 'GFI': '> .90 (good)', 'AGFI': '> .85 (good)',
    'NFI': '> .90 (good)', 'TLI': '> .95 (good)',
    'RMSEA': '< .05 (good), < .08 (acceptable)',
    'AIC': 'lower = better', 'BIC': 'lower = better'
}
for r_i, (idx_name, thresh) in enumerate(thresholds.items(), 4):
    if idx_name in fit_nr.index:
        val = fit_nr.loc[idx_name, 'Value']
        ws3.cell(r_i, 1, idx_name).alignment = Alignment(horizontal='left')
        ws3.cell(r_i, 2, round(float(val), 4)).alignment = Alignment(horizontal='center')
        ws3.cell(r_i, 3, thresh).alignment = Alignment(horizontal='left')
border_range(ws3, 3, 1, 3 + len(thresholds), 3)

# ── Sheet 4: 효과 분해 ───────────────────────────────────────
ws4 = wb.create_sheet('Effect_Decomp')
ws4.column_dimensions['A'].width = 50
ws4.column_dimensions['B'].width = 15

ws4['A1'] = 'Effect Decomposition'
ws4['A1'].font = Font(bold=True, size=12)
ws4.merge_cells('A1:B1')
for i, h in enumerate(['Effect Type', 'Value'], 1):
    ws4.cell(3, i, h)
style_header(ws4, 3, 2, color='4B0082')

effect_data = [
    ('=== urban_disparity -> carbon ===',              ''),
    ('  Direct Effect  (urban_disparity -> carbon)',   safe_round(c)),
    ('  Path a  (urban_disparity -> UCI)',              safe_round(a)),
    ('  Path b  (UCI -> carbon)',                       safe_round(b)),
    ('  Indirect Effect  (a x b)',                      safe_round(indirect)),
    ('  Total Effect  (direct + indirect)',             safe_round(total)),
    ('  Reverse path  (UCI -> urban_disparity)',        safe_round(d)),
    ('',                                               ''),
    ('=== urban_dev -> carbon (Control) ===',          ''),
    ('  Direct Effect  (urban_dev -> carbon)',          safe_round(e_dev_carbon)),
    ('  Indirect via UCI  (urban_dev -> UCI -> carbon)', safe_round(dev_indirect)),
    ('  Total Effect',                                 safe_round(dev_total)),
]
for r_i, (label, val) in enumerate(effect_data, 4):
    ws4.cell(r_i, 1, label)
    if val != '':
        ws4.cell(r_i, 2, val).alignment = Alignment(horizontal='center')
border_range(ws4, 3, 1, 3 + len(effect_data), 2)

# ── Sheet 5: 2SLS 결과 ───────────────────────────────────────
ws5 = wb.create_sheet('2SLS_Results')
for col, w in zip(['A','B','C','D','E','F','G'], [22,18,13,13,13,13,8]):
    ws5.column_dimensions[col].width = w

ws5['A1'] = '2SLS Results (IV Estimation, Robust SE)'
ws5['A1'].font = Font(bold=True, size=12)
ws5.merge_cells('A1:G1')
ws5['A2'] = (
    'Endogenous: uci_score, disparity_score  |  '
    'Exogenous control: dev_score  |  IV: UCI + Disparity indicators'
)
ws5['A2'].font = Font(italic=True, size=9, color='555555')
ws5.merge_cells('A2:G2')

h2 = ['Dependent', 'Variable', 'Coef.', 'Std. Err', 'T-stat', 'P-value', 'Sig.']
for i, h in enumerate(h2, 1):
    ws5.cell(4, i, h)
style_header(ws5, 4, len(h2), color='8B0000')

row_idx = 5
carbon_labels = {
    'build_elec_e_pc': 'Building Electricity',
    'transport_e_pc':  'Transport Energy',
}
var_labels = {
    'uci_score':       'UCI Score (endog.)',
    'disparity_score': 'Disparity Score (endog.)',
    'dev_score':       'Dev Score (exog. control)'
}

for cv, res in iv_results.items():
    params_iv = res.params
    se_iv     = res.std_errors
    tstat_iv  = res.tstats
    pval_iv   = res.pvalues
    first_row = True
    for var in ['uci_score', 'disparity_score', 'dev_score']:
        dep_label = carbon_labels[cv] if first_row else ''
        p = pval_iv[var]
        ws5.cell(row_idx, 1, dep_label)
        ws5.cell(row_idx, 2, var_labels[var])
        ws5.cell(row_idx, 3, round(params_iv[var], 4)).alignment = Alignment(horizontal='center')
        ws5.cell(row_idx, 4, round(se_iv[var], 4)).alignment     = Alignment(horizontal='center')
        ws5.cell(row_idx, 5, round(tstat_iv[var], 3)).alignment  = Alignment(horizontal='center')
        ws5.cell(row_idx, 6, round(p, 4)).alignment              = Alignment(horizontal='center')
        ws5.cell(row_idx, 7, sig_star(p)).alignment              = Alignment(horizontal='center')
        first_row = False
        row_idx += 1
    row_idx += 1  # 그룹 간 빈 행

border_range(ws5, 4, 1, row_idx - 2, len(h2))

wb.save(f'{OUT}/SEM_2SLS_Tables.xlsx')
print("엑셀 저장 완료")


# ════════════════════════════════════════════════════════════════
# 그래프 1: Path Diagram — 박스-화살표 겹침 방지 재설계
# 전략: 박스 간 간격 충분히 확보, 화살표는 박스 외부 공간만 통과
# ════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(20, 13))
ax.set_xlim(0, 20); ax.set_ylim(0, 13); ax.axis('off')
fig.patch.set_facecolor('#FFFFFF')

def draw_box(ax, x, y, w, h, lines, color='#2C5F8A', text_color='white'):
    box = mpatches.FancyBboxPatch(
        (x - w/2, y - h/2), w, h,
        boxstyle='round,pad=0.2', linewidth=2.5,
        edgecolor='white', facecolor=color, zorder=5
    )
    ax.add_patch(box)
    if isinstance(lines, str):
        lines = [lines]
    total_h = len(lines) * 0.42
    for i, line in enumerate(lines):
        ty = y + total_h/2 - 0.42*i - 0.21
        fs = 13 if i == 0 else 9.5
        fw = 'bold' if i == 0 else 'normal'
        ax.text(x, ty, line, ha='center', va='center',
                fontsize=fs, fontweight=fw, color=text_color, zorder=6)

def draw_arrow(ax, x1, y1, x2, y2, label='', color='#333333', lw=2.5,
               curvature=0.0, label_offset=(0, 0), linestyle='-',
               label_fontsize=12, zorder=4):
    arrow = FancyArrowPatch(
        (x1, y1), (x2, y2),
        arrowstyle='->,head_length=0.6,head_width=0.35',
        color=color,
        connectionstyle=f'arc3,rad={curvature}',
        linewidth=lw, mutation_scale=22, zorder=zorder,
        linestyle=linestyle
    )
    ax.add_patch(arrow)
    if label:
        mx = (x1 + x2)/2 + label_offset[0]
        my = (y1 + y2)/2 + label_offset[1]
        ax.text(mx, my, label, ha='center', va='center',
                fontsize=label_fontsize,
                color=color, fontweight='bold', zorder=zorder + 1,
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                          alpha=0.97, edgecolor=color, linewidth=1.2))

# ── 박스 배치 (넓은 간격) ─────────────────────────────────────
# main row y=8.5, 박스 간 2.5+ 유닛 간격
BH = 1.7

# Disparity: center x=3, 폭 4.5 → 좌0.75 ~ 우5.25
draw_box(ax, 3.0, 8.5, 4.5, BH,
         ['Urban Disparity', f'({", ".join(vars_disparity)})'],
         color='#1A5276')

# UCI: center x=10, 폭 3.2 → 좌8.4 ~ 우11.6
draw_box(ax, 10.0, 8.5, 3.2, BH,
         ['UCI', f'({", ".join(vars_uci)})'],
         color='#117A65')

# Carbon: center x=17, 폭 4.5 → 좌14.75 ~ 우19.25
draw_box(ax, 17.0, 8.5, 4.5, BH,
         ['Carbon Emissions', f'({", ".join(vars_carbon)})'],
         color='#7B241C')

# Dev: center x=10, y=2.5
draw_box(ax, 10.0, 2.5, 4.5, BH,
         ['Urban Dev Level', f'({", ".join(vars_dev)})', '(Control)'],
         color='#7D6608')

# ── 주요 구조경로 (실선, 굵게, 박스 밖으로만) ────────────────

# disparity → UCI : 상단 직선 (박스 위쪽 가장자리 바로 위)
draw_arrow(ax, 5.25, 9.0, 8.4, 9.0,
           f'β = {a:.3f}{sig_star(a_p)}',
           color='#0E6251', lw=3.0, curvature=0.0,
           label_offset=(0, 0.55), label_fontsize=12)

# UCI → disparity : 하단 직선 (박스 아래쪽 가장자리 바로 아래)
draw_arrow(ax, 8.4, 8.0, 5.25, 8.0,
           f'β = {d:.3f}{sig_star(d_p)}',
           color='#1A5276', lw=3.0, curvature=0.0,
           label_offset=(0, -0.55), label_fontsize=12)

# UCI → carbon : 중앙 직선
draw_arrow(ax, 11.6, 8.5, 14.75, 8.5,
           f'β = {b:.3f}{sig_star(b_p)}',
           color='#922B21', lw=3.0, curvature=0.0,
           label_offset=(0, 0.55), label_fontsize=12)

# disparity → carbon (직접효과) : 큰 아치 — 모든 박스 위를 지남
draw_arrow(ax, 5.25, 9.35, 14.75, 9.35,
           f"β = {c:.3f}{sig_star(c_p)}  (direct)",
           color='#7D3C98', lw=2.5, curvature=-0.18,
           label_offset=(0, 1.3), label_fontsize=12)

# ── 통제변수 화살표 (점선) — 모두 아래쪽 빈 공간을 통과 ─────

# urban_dev → disparity : 좌상 대각선
draw_arrow(ax, 8.2, 3.35, 4.2, 7.65,
           f'β = {e_dev_disp:.3f}{sig_star(e_dev_disp_p)}',
           color='#7D6608', lw=2.2, curvature=0.0, linestyle='--',
           label_offset=(-1.0, 0.3), label_fontsize=11)

# urban_dev → UCI : 수직 위
draw_arrow(ax, 10.0, 3.35, 10.0, 7.65,
           f'β = {e_dev_uci:.3f}{sig_star(e_dev_uci_p)}',
           color='#7D6608', lw=2.2, curvature=0.0, linestyle='--',
           label_offset=(1.0, 0), label_fontsize=11)

# urban_dev → carbon : 우상 대각선
draw_arrow(ax, 11.8, 3.35, 15.8, 7.65,
           f'β = {e_dev_carbon:.3f}{sig_star(e_dev_carb_p)}',
           color='#7D6608', lw=2.2, curvature=0.0, linestyle='--',
           label_offset=(1.0, 0.3), label_fontsize=11)

# ── 효과 분해 박스 (좌하단) ──────────────────────────────────
ax.text(3.5, 2.5,
        f'Effect Decomposition\n'
        f'urban_disparity → carbon\n'
        f'─────────────────────\n'
        f'Indirect (a×b): {a:.3f} × {b:.3f} = {indirect:.3f}\n'
        f"Direct   (c') : {c:.3f}\n"
        f'Total         : {total:.3f}',
        ha='center', va='center', fontsize=11, family='monospace',
        bbox=dict(boxstyle='round,pad=0.5', facecolor='#F0F4F8',
                  edgecolor='#1A5276', linewidth=1.5))

# ── 범례 (우하단) ────────────────────────────────────────────
legend_elements = [
    mpatches.Patch(facecolor='#1A5276', label='Urban Disparity'),
    mpatches.Patch(facecolor='#117A65', label='UCI'),
    mpatches.Patch(facecolor='#7B241C', label='Carbon Emissions'),
    mpatches.Patch(facecolor='#7D6608', label='Urban Dev (Control)'),
    plt.Line2D([0], [0], color='#333', lw=3, label='Structural paths'),
    plt.Line2D([0], [0], color='#7D6608', lw=2, linestyle='--', label='Control paths'),
]
ax.legend(handles=legend_elements, loc='lower right', fontsize=10,
          framealpha=0.95, ncol=2, edgecolor='#CCCCCC',
          bbox_to_anchor=(0.98, 0.02))

ax.set_title(
    'Non-recursive SEM Path Diagram\n'
    'Urban Disparity ↔ UCI → Carbon Emissions  |  Urban Dev Level (Control)',
    fontsize=15, fontweight='bold', pad=16
)
ax.text(0.01, 0.01, '*** p < .001    ** p < .01    * p < .05',
        transform=ax.transAxes, fontsize=9.5, color='#555555')

plt.savefig(f'{OUT}/Fig1_PathDiagram.png', dpi=200, bbox_inches='tight')
plt.close()
print("경로도 저장 완료")


# ════════════════════════════════════════════════════════════════
# 그래프 2: 2SLS 계수 비교 (탄소변수별, 3개 변수)
# ════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(1, 2, figsize=(11, 5.5), sharey=False)
fig.patch.set_facecolor('#F8F9FA')

carbon_full = {
    'build_elec_e_pc': 'Building\nElectricity',
    'transport_e_pc':  'Transport\nEnergy',
}
colors_iv = {
    'uci_score':       '#117A65',
    'disparity_score': '#1A5276',
    'dev_score':       '#7D6608'
}
var_names_iv = {
    'uci_score':       'UCI',
    'disparity_score': 'Disparity',
    'dev_score':       'Dev (ctrl)'
}

for ax, (cv, res) in zip(axes, iv_results.items()):
    params_iv = res.params
    ci        = res.conf_int()
    pval_iv   = res.pvalues

    vars_plot = ['uci_score', 'disparity_score', 'dev_score']
    coefs  = [params_iv[v] for v in vars_plot]
    lowers = [params_iv[v] - ci.loc[v, 'lower'] for v in vars_plot]
    uppers = [ci.loc[v, 'upper'] - params_iv[v] for v in vars_plot]
    clrs   = [colors_iv[v] for v in vars_plot]
    xlabels = [var_names_iv[v] for v in vars_plot]

    bars = ax.bar(xlabels, coefs, color=clrs, width=0.5,
                  edgecolor='white', linewidth=1.2, zorder=3)
    ax.errorbar(xlabels, coefs, yerr=[lowers, uppers], fmt='none',
                color='#333333', capsize=6, linewidth=1.5, zorder=4)

    for bar, v, coef in zip(bars, vars_plot, coefs):
        p    = pval_iv[v]
        star = sig_star(p)
        idx  = vars_plot.index(v)
        ypos = coef + (uppers[idx] + 0.03) if coef >= 0 else coef - (lowers[idx] + 0.03)
        ax.text(bar.get_x() + bar.get_width()/2, ypos,
                f'{coef:.3f}{star}',
                ha='center', va='bottom' if coef >= 0 else 'top',
                fontsize=9.5, fontweight='bold', color='#1B2631')

    ax.axhline(0, color='#888888', linewidth=0.8, linestyle='--')
    ax.set_title(carbon_full[cv], fontsize=11, fontweight='bold')
    ax.set_ylabel('2SLS Coefficient' if ax == axes[0] else '', fontsize=9)
    ax.set_facecolor('#F8F9FA')
    ax.spines[['top', 'right']].set_visible(False)
    ax.grid(axis='y', alpha=0.3, zorder=0)
    ymin, ymax = ax.get_ylim()
    pad = (ymax - ymin) * 0.28
    ax.set_ylim(ymin - pad, ymax + pad)

fig.suptitle('2SLS Estimates by Carbon Emission Variable\n(Robust SE, 95% CI)',
             fontsize=13, fontweight='bold', y=1.01)

legend_patches = [
    mpatches.Patch(color='#117A65', label='UCI Score (endog.)'),
    mpatches.Patch(color='#1A5276', label='Disparity Score (endog.)'),
    mpatches.Patch(color='#7D6608', label='Dev Score (exog. control)'),
]
fig.legend(handles=legend_patches, loc='lower center', ncol=3, fontsize=9.5,
           bbox_to_anchor=(0.5, -0.06), framealpha=0.9)

plt.tight_layout()
plt.savefig(f'{OUT}/Fig2_2SLS_Coefficients.png', dpi=180, bbox_inches='tight')
plt.close()
print("2SLS 계수도 저장 완료")


# ════════════════════════════════════════════════════════════════
# 그래프 3: SEM 측정모델 Factor Loadings (2×2 레이아웃)
# ════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(2, 2, figsize=(13, 9))
fig.patch.set_facecolor('#F8F9FA')

constructs = [
    ('urban_disparity', vars_disparity, '#1A5276', 'Urban Disparity'),
    ('urban_dev',       vars_dev,       '#7D6608', 'Urban Dev Level (Control)'),
    ('UCI',             vars_uci,       '#117A65', 'UCI'),
    ('carbon',          vars_carbon,    '#7B241C', 'Carbon'),
]

meas_rows = params_nr[measurement_mask].copy()

for ax, (latent, indicators, color, title) in zip(axes.flat, constructs):
    rows = meas_rows[meas_rows['rval'] == latent]
    loadings, pvals = [], []
    for ind in indicators:
        r = rows[rows['lval'] == ind]
        loadings.append(r['Estimate'].values[0] if len(r) else np.nan)
        pvals.append(r['p-value'].values[0] if len(r) else np.nan)

    def is_sig(p):
        try: return float(p) < 0.05
        except: return False

    bar_colors = [color if is_sig(p) else '#AAAAAA' for p in pvals]
    bars = ax.barh(indicators, loadings, color=bar_colors,
                   edgecolor='white', height=0.45, zorder=3)

    xmin, xmax = ax.get_xlim()
    x_range = xmax - xmin
    ax.set_xlim(xmin - x_range * 0.05, xmax + x_range * 0.28)

    for bar, lv, p in zip(bars, loadings, pvals):
        if pd.isna(lv): continue
        star = sig_star(p)
        by = bar.get_y() + bar.get_height() / 2
        if lv >= 0:
            ax.text(lv + 0.02, by, f'{lv:.3f}{star}',
                    va='center', ha='left', fontsize=8.5, color='#1B2631')
        else:
            ax.text(0.02, by, f'{lv:.3f}{star}',
                    va='center', ha='left', fontsize=8.5, color='#1B2631')

    ax.axvline(0, color='#888888', linewidth=0.8, linestyle='--')
    ax.set_title(title, fontsize=10.5, fontweight='bold', color=color)
    ax.set_xlabel('Factor Loading', fontsize=8.5)
    ax.set_facecolor('#F8F9FA')
    ax.spines[['top', 'right']].set_visible(False)
    ax.grid(axis='x', alpha=0.3, zorder=0)
    ax.tick_params(axis='y', labelsize=9, pad=4)

fig.suptitle('SEM Measurement Model — Factor Loadings  (gray = p ≥ .05)',
             fontsize=13, fontweight='bold')
plt.tight_layout(pad=1.8)
plt.savefig(f'{OUT}/Fig3_FactorLoadings.png', dpi=180, bbox_inches='tight')
plt.close()
print("측정모델도 저장 완료")


# ════════════════════════════════════════════════════════════════
# 그래프 4: 효과 분해 (urban_disparity + urban_dev -> carbon)
# ════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(1, 2, figsize=(12, 4.5))
fig.patch.set_facecolor('#F8F9FA')

# 왼쪽: urban_disparity -> carbon
ax = axes[0]
effect_labels = ["Direct\n(c')", 'Indirect\n(a×b)', 'Total']
effect_vals   = [c, indirect, total]
eff_colors    = ['#884EA0', '#117A65', '#2C3E50']

bars = ax.bar(effect_labels, effect_vals, color=eff_colors, width=0.45,
              edgecolor='white', linewidth=1.2, zorder=3)
ax.axhline(0, color='#888888', linewidth=0.8, linestyle='--')
for bar, val in zip(bars, effect_vals):
    ypos = val + 0.012 if val >= 0 else val - 0.018
    ax.text(bar.get_x() + bar.get_width()/2, ypos, f'{val:.4f}',
            ha='center', va='bottom' if val >= 0 else 'top',
            fontsize=10.5, fontweight='bold', color='#1B2631')
ymin, ymax = ax.get_ylim()
ax.set_ylim(ymin - (ymax - ymin)*0.25, ymax + (ymax - ymin)*0.25)
ax.set_title('urban_disparity → carbon\n(via UCI)', fontsize=11, fontweight='bold')
ax.set_ylabel('Standardized Effect', fontsize=9)
ax.set_facecolor('#F8F9FA')
ax.spines[['top', 'right']].set_visible(False)
ax.grid(axis='y', alpha=0.3, zorder=0)

# 오른쪽: urban_dev -> carbon (통제변수)
ax = axes[1]
effect_labels2 = ['Direct\n(dev→carb)', 'Indirect\n(dev→UCI→carb)', 'Total']
effect_vals2   = [e_dev_carbon, dev_indirect, dev_total]
eff_colors2    = ['#9A7D0A', '#7D6608', '#4A4A00']

bars2 = ax.bar(effect_labels2, effect_vals2, color=eff_colors2, width=0.45,
               edgecolor='white', linewidth=1.2, zorder=3)
ax.axhline(0, color='#888888', linewidth=0.8, linestyle='--')
for bar, val in zip(bars2, effect_vals2):
    ypos = val + 0.012 if val >= 0 else val - 0.018
    ax.text(bar.get_x() + bar.get_width()/2, ypos, f'{val:.4f}',
            ha='center', va='bottom' if val >= 0 else 'top',
            fontsize=10.5, fontweight='bold', color='#1B2631')
ymin, ymax = ax.get_ylim()
ax.set_ylim(ymin - (ymax - ymin)*0.25, ymax + (ymax - ymin)*0.25)
ax.set_title('urban_dev → carbon\n(Control Variable)', fontsize=11, fontweight='bold', color='#7D6608')
ax.set_ylabel('Standardized Effect', fontsize=9)
ax.set_facecolor('#F8F9FA')
ax.spines[['top', 'right']].set_visible(False)
ax.grid(axis='y', alpha=0.3, zorder=0)

fig.suptitle('Effect Decomposition: Structural Paths to Carbon Emissions',
             fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUT}/Fig4_EffectDecomp.png', dpi=180, bbox_inches='tight')
plt.close()
print("효과분해도 저장 완료")


# ════════════════════════════════════════════════════════════════
# 그래프 5: 개별 관측변수 기여도 (4개 구성개념)
# ════════════════════════════════════════════════════════════════
meas_rows_all = params_nr[measurement_mask].copy()

contrib_data = []
for ind in vars_disparity:
    r  = meas_rows_all[meas_rows_all['lval'] == ind]
    lv = r['Estimate'].values[0] if len(r) else np.nan
    contrib_data.append({'variable': ind, 'group': 'Urban Disparity',
                         'loading': lv, 'total_contrib': lv * total})
for ind in vars_dev:
    r  = meas_rows_all[meas_rows_all['lval'] == ind]
    lv = r['Estimate'].values[0] if len(r) else np.nan
    contrib_data.append({'variable': ind, 'group': 'Urban Dev (Control)',
                         'loading': lv, 'total_contrib': lv * dev_total})
for ind in vars_uci:
    r  = meas_rows_all[meas_rows_all['lval'] == ind]
    lv = r['Estimate'].values[0] if len(r) else np.nan
    contrib_data.append({'variable': ind, 'group': 'UCI',
                         'loading': lv, 'total_contrib': lv * b})
for ind in vars_carbon:
    r  = meas_rows_all[meas_rows_all['lval'] == ind]
    lv = r['Estimate'].values[0] if len(r) else np.nan
    contrib_data.append({'variable': ind, 'group': 'Carbon',
                         'loading': lv, 'total_contrib': lv})

df_contrib = pd.DataFrame(contrib_data)

# Sheet 6: 기여도 엑셀
ws6 = wb.create_sheet('Indicator_Contribution')
for col, w in zip(['A','B','C','D'], [22, 22, 14, 18]):
    ws6.column_dimensions[col].width = w
ws6['A1'] = 'Indicator Variable Contribution to Carbon (via Latent Paths)'
ws6['A1'].font = Font(bold=True, size=12)
ws6.merge_cells('A1:D1')
ws6['A2'] = 'Total Contribution = Factor Loading x Latent->Carbon Total Effect'
ws6['A2'].font = Font(italic=True, size=9, color='555555')
ws6.merge_cells('A2:D2')
for i, h in enumerate(['Variable', 'Group', 'Factor Loading', 'Total Contribution'], 1):
    ws6.cell(4, i, h)
style_header(ws6, 4, 4, color='4A235A')
for r_i, row in df_contrib.iterrows():
    ws6.cell(r_i + 5, 1, row['variable'])
    ws6.cell(r_i + 5, 2, row['group'])
    ws6.cell(r_i + 5, 3, safe_round(row['loading'])).alignment   = Alignment(horizontal='center')
    ws6.cell(r_i + 5, 4, safe_round(row['total_contrib'])).alignment = Alignment(horizontal='center')
border_range(ws6, 4, 1, len(df_contrib) + 5, 4)
wb.save(f'{OUT}/SEM_2SLS_Tables.xlsx')

# 그래프: 2x2 패널
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.patch.set_facecolor('#F8F9FA')

groups_plot = [
    ('Urban Disparity',    '#1A5276', 'loading'),
    ('Urban Dev (Control)','#7D6608', 'loading'),
    ('UCI',                '#117A65', 'loading'),
    ('Carbon',             '#7B241C', 'loading'),
]

for ax, (grp, color, col) in zip(axes.flat, groups_plot):
    sub  = df_contrib[df_contrib['group'] == grp].copy()
    vals = sub[col].values
    labels = sub['variable'].values
    bar_colors = [color if not pd.isna(v) and v != 0 else '#CCCCCC' for v in vals]
    bars = ax.barh(labels, vals, color=bar_colors,
                   edgecolor='white', height=0.42, zorder=3)

    xmin, xmax = ax.get_xlim()
    x_range = max(abs(xmin), abs(xmax), 0.1)
    ax.set_xlim(xmin - x_range*0.05, xmax + x_range*0.3)

    for bar, v in zip(bars, vals):
        if pd.isna(v): continue
        by = bar.get_y() + bar.get_height() / 2
        if v >= 0:
            ax.text(v + 0.01, by, f'{v:.3f}',
                    va='center', ha='left', fontsize=8.5, color='#1B2631')
        else:
            ax.text(0.01, by, f'{v:.3f}',
                    va='center', ha='left', fontsize=8.5, color='#1B2631')

    ax.axvline(0, color='#888888', linewidth=0.8, linestyle='--')
    ax.set_title(grp, fontsize=10.5, fontweight='bold', color=color)
    ax.set_xlabel('Factor Loading (onto latent)', fontsize=8.5)
    ax.set_facecolor('#F8F9FA')
    ax.spines[['top', 'right']].set_visible(False)
    ax.grid(axis='x', alpha=0.3, zorder=0)
    ax.tick_params(axis='y', labelsize=9, pad=5)

fig.suptitle('Individual Variable Factor Loadings by Construct',
             fontsize=13, fontweight='bold')
plt.tight_layout(pad=1.8)
plt.savefig(f'{OUT}/Fig5_IndicatorContribution.png', dpi=180, bbox_inches='tight')
plt.close()
print("개별 변수 기여도 저장 완료")


# ════════════════════════════════════════════════════════════════
# ZIP 묶기
# ════════════════════════════════════════════════════════════════
zip_path = 'SEM_PathAnalysis_Results#6.zip'
with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
    for fname in os.listdir(OUT):
        zf.write(os.path.join(OUT, fname), fname)

print(f"\n완료: {zip_path}")
print("포함 파일:")
with zipfile.ZipFile(zip_path) as zf:
    for name in zf.namelist():
        print(f"  {name}")
